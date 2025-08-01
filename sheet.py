import os
import datetime
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from openpyxl import load_workbook, Workbook
from mtcnn.mtcnn import MTCNN
import cv2
import time

# Main GUI setup
root = tk.Tk()
root.title("Face Detection & Attendance System")
root.geometry("750x450")
root.config(bg="#f0f0f0")

# GUI State Variables
output_path_var = tk.StringVar()
resolution_var = tk.StringVar(value="640x480")
gpu_frac_var = tk.DoubleVar(value=0.5)
face_size_var = tk.StringVar(value="160x160")
user_name_var = tk.StringVar()
dataset_source_var = tk.StringVar(value="Webcam")
video_path_var = tk.StringVar()

def mark_present(st_name):
    names = os.listdir('output/')
    print("Names found in output folder:", names)

    sub = 'SAMPLE'
    filepath = os.path.join('attendance', f'{sub}.xlsx')

    if not os.path.exists('attendance'):
        os.makedirs('attendance')

    if not os.path.exists(filepath):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = sub

        sheet.cell(row=1, column=1, value="Name")
        sheet.cell(row=1, column=2, value=str(datetime.datetime.now().date()))

        row = 2
        for name in names:
            sheet.cell(row=row, column=1, value=name)
            row += 1

        workbook.save(filepath)
        print("Created new attendance file:", filepath)

    workbook = load_workbook(filepath)
    sheet = workbook.active

    col = sheet.max_column + 1
    sheet.cell(row=1, column=col, value=str(datetime.datetime.now().date()))

    row = 2
    for name in names:
        mark = 'P' if st_name in name else 'A'
        sheet.cell(row=row, column=col, value=mark)
        row += 1

    workbook.save(filepath)
    print("Attendance marked for:", st_name)

def browse_output_path():
    folder = filedialog.askdirectory()
    output_path_var.set(folder)

def on_continue():
    name = user_name_var.get().strip()
    resolution = resolution_var.get()
    dataset_mode = dataset_source_var.get()

    if not name:
        messagebox.showerror("Missing Name", "Please enter your name to continue.")
        return

    width, height = map(int, resolution.split('x'))
    save_path = os.path.join("output", name)
    os.makedirs(save_path, exist_ok=True)

    video_filename = os.path.join(save_path, f"{name}.avi")
    cap = cv2.VideoCapture(0)
    cap.set(3, width)
    cap.set(4, height)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter(video_filename, fourcc, 20.0, (width, height))

    messagebox.showinfo("Recording", f"Recording 5 seconds of video for {name}. Please stay still.")

    start_time = time.time()
    while int(time.time() - start_time) < 5:
        ret, frame = cap.read()
        if not ret:
            break
        out.write(frame)
        cv2.imshow("Recording... Press 'q' to abort", frame)
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    out.release()
    cv2.destroyAllWindows()

    # Process saved video to extract faces
    cap = cv2.VideoCapture(video_filename)
    cap.set(3, 320)
    cap.set(4, 240)
    detector = MTCNN()
    count = 0
    max_images = 50

    while cap.isOpened():
        ret, frame = cap.read()
        if not ret or count >= max_images:
            break

        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)  # Convert to RGB for MTCNN
        faces = detector.detect_faces(rgb_frame)
        print(f"Detected {len(faces)} faces in frame")

        for face in faces:
            if 'box' in face:
                x, y, w, h = face['box']
                x, y = max(0, x), max(0, y)
                face_crop = frame[y:y+h, x:x+w]

                try:
                    face_resized = cv2.resize(face_crop, (160, 160))
                    filename = os.path.join(save_path, f"{name}_{count}.jpg")
                    cv2.imwrite(filename, face_resized)
                    count += 1
                    print(f"Saved: {filename}")
                except Exception as e:
                    print(f"Skipping one face due to error: {e}")

            if count >= max_images:
                break

    cap.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Done", f"Captured {count} face images from video for {name}.")
    cap.release()
    cv2.destroyAllWindows()
    messagebox.showinfo("Done", f"Captured {count} face images from video for {name}.")

# Layout
row = 0

# Output Path
tk.Label(root, text="Select output folder:").grid(row=row, column=0, sticky="w", padx=10, pady=5)
tk.Entry(root, textvariable=output_path_var, width=50).grid(row=row, column=1, padx=5)
tk.Button(root, text="Browse", command=browse_output_path).grid(row=row, column=2)
row += 1

# Webcam Resolution
tk.Label(root, text="Webcam Resolution:").grid(row=row, column=0, sticky="w", padx=10, pady=5)
ttk.Combobox(root, textvariable=resolution_var, values=["640x480", "1280x720"]).grid(row=row, column=1, padx=5)
row += 1

# GPU Fraction
tk.Label(root, text="GPU Memory Fraction (0 to 1):").grid(row=row, column=0, sticky="w", padx=10, pady=5)
tk.Scale(root, from_=0.0, to=1.0, resolution=0.1, orient="horizontal", variable=gpu_frac_var).grid(row=row, column=1, sticky="we", padx=5)
row += 1

# Face Size
tk.Label(root, text="Face Size (WxH):").grid(row=row, column=0, sticky="w", padx=10, pady=5)
ttk.Combobox(root, textvariable=face_size_var, values=["160x160", "224x224", "96x96"]).grid(row=row, column=1, padx=5)
row += 1

# Name Entry
tk.Label(root, text="Enter Your Name:").grid(row=row, column=0, sticky="w", padx=10, pady=5)
tk.Entry(root, textvariable=user_name_var).grid(row=row, column=1, padx=5)
row += 1

# Dataset Source (Radio)
tk.Label(root, text="Create Dataset Using:").grid(row=row, column=0, sticky="w", padx=10, pady=5)
tk.Radiobutton(root, text="Video", variable=dataset_source_var, value="Video").grid(row=row, column=1, sticky="w")
tk.Radiobutton(root, text="Webcam", variable=dataset_source_var, value="Webcam").grid(row=row, column=1, sticky="e")
row += 1

# Video Path (Optional)
tk.Label(root, text="Enter video path (if applicable):").grid(row=row, column=0, sticky="w", padx=10, pady=5)
tk.Entry(root, textvariable=video_path_var, width=50).grid(row=row, column=1, padx=5)
row += 1

# Buttons
tk.Button(root, text="CONTINUE", bg="cyan", command=on_continue).grid(row=row, column=1, pady=20)
tk.Button(root, text="EXIT", bg="tomato", command=root.quit).grid(row=row, column=2, pady=20)

root.mainloop()