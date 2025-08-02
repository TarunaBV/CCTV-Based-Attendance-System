import os
import cv2
import time
import datetime
import numpy as np
import streamlit as st
from mtcnn.mtcnn import MTCNN
from keras_facenet import FaceNet
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import load_workbook, Workbook
import pickle

# Paths
output_folder = "output"
attendance_folder = "attendance"
embedding_file = "face_embeddings.pkl"
os.makedirs(output_folder, exist_ok=True)
os.makedirs(attendance_folder, exist_ok=True)

# Initialize detector and embedder
detector = MTCNN()
embedder = FaceNet()

# Load embeddings if they exist
if os.path.exists(embedding_file):
    with open(embedding_file, "rb") as f:
        known_faces = pickle.load(f)
else:
    known_faces = []  # List of {'name': ..., 'embedding': ...}

# Streamlit UI
st.set_page_config(page_title="Face Dataset Capture", layout="centered")
st.title("📸 Face Dataset Capture & Attendance")
st.markdown("Create your face dataset using webcam and mark attendance.")

# Input fields
name = st.text_input("👤 Enter your name")
resolution = st.selectbox("🎥 Select resolution", ["640x480", "1280x720"])
face_size = st.selectbox("📐 Face image size", ["160x160", "224x224", "96x96"])
capture_duration = st.slider("⏱️ Capture duration (sec)", 3, 10, 5)


def get_face_embedding(image):
    rgb_img = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
    faces = detector.detect_faces(rgb_img)
    if not faces:
        return None
    x, y, w, h = faces[0]['box']
    x, y = max(0, x), max(0, y)
    face = rgb_img[y:y+h, x:x+w]
    face = cv2.resize(face, (160, 160))  # FaceNet input size
    embedding = embedder.embeddings([face])[0]
    return embedding


def capture_faces_from_webcam(name, resolution, face_size, duration):
    width, height = map(int, resolution.split("x"))
    w_face, h_face = map(int, face_size.split("x"))
    save_path = os.path.join(output_folder, name)
    os.makedirs(save_path, exist_ok=True)

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    video_filename = os.path.join(save_path, f"{name}_{timestamp}.avi")
    cap = cv2.VideoCapture(0)
    cap.set(3, width)
    cap.set(4, height)
    fourcc = cv2.VideoWriter_fourcc(*'XVID')
    out = cv2.VideoWriter(video_filename, fourcc, 20.0, (width, height))

    st.info("Recording video for face capture...")
    start_time = time.time()
    while int(time.time() - start_time) < duration:
        ret, frame = cap.read()
        if not ret:
            break
        out.write(frame)
        frame_resized = cv2.resize(frame, (320, 240))
        frame_bgr = cv2.cvtColor(frame_resized, cv2.COLOR_BGR2RGB)
        st.image(frame_bgr, channels="RGB", caption="Recording...", use_container_width=True)

    cap.release()
    out.release()
    return video_filename, save_path, (w_face, h_face)


def extract_faces(video_path, save_path, face_dim):
    cap = cv2.VideoCapture(video_path)
    count = 0
    max_images = 50

    st.info("Detecting and saving faces from video...")
    while cap.isOpened() and count < max_images:
        ret, frame = cap.read()
        if not ret:
            break
        rgb_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        faces = detector.detect_faces(rgb_frame)

        for face in faces:
            if 'box' in face:
                x, y, w, h = face['box']
                x, y = max(0, x), max(0, y)
                face_crop = frame[y:y+h, x:x+w]
                try:
                    face_resized = cv2.resize(face_crop, face_dim)
                    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = os.path.join(save_path, f"{name}_{timestamp}_{count}.jpg")
                    cv2.imwrite(filename, face_resized)
                    count += 1
                except:
                    continue
        if count >= max_images:
            break
    cap.release()
    return count


def mark_attendance_excel(name):
    today = str(datetime.datetime.now().date())
    filepath = os.path.join(attendance_folder, 'SAMPLE.xlsx')

    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = 'SAMPLE'
        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value=today)
        wb.save(filepath)

    wb = load_workbook(filepath)
    ws = wb.active
    col = ws.max_column + 1 if ws.cell(row=1, column=ws.max_column).value != today else ws.max_column

    if ws.cell(row=1, column=col).value != today:
        ws.cell(row=1, column=col, value=today)

    existing_names = [ws.cell(row=i, column=1).value for i in range(2, ws.max_row + 1)]
    if name in existing_names:
        idx = existing_names.index(name) + 2
    else:
        idx = ws.max_row + 1
        ws.cell(row=idx, column=1, value=name)

    ws.cell(row=idx, column=col, value="P")
    wb.save(filepath)
    return filepath


if st.button("▶️ Start Capture"):
    if not name:
        st.error("Please enter your name.")
    else:
        # Step 1: Capture one frame to check for duplicates
        temp_cap = cv2.VideoCapture(0)
        ret, frame = temp_cap.read()
        temp_cap.release()

        if not ret:
            st.error("Unable to access camera.")
        else:
            new_embedding = get_face_embedding(frame)
            if new_embedding is None:
                st.warning("No face detected. Make sure your face is clearly visible.")
            else:
                duplicate = False
                for record in known_faces:
                    similarity = cosine_similarity([new_embedding], [record['embedding']])[0][0]
                    if similarity > 0.6:
                        duplicate = True
                        break

                if duplicate:
                    st.warning("This face has already been captured. Duplicate entry blocked.")
                else:
                    video_path, save_path, face_dim = capture_faces_from_webcam(name, resolution, face_size, capture_duration)
                    count = extract_faces(video_path, save_path, face_dim)

                    if count > 0:
                        known_faces.append({'name': name, 'embedding': new_embedding})
                        with open(embedding_file, "wb") as f:
                            pickle.dump(known_faces, f)

                        att_path = mark_attendance_excel(name)
                        st.success(f"✅ Captured {count} face images for {name}")
                        st.info(f"Attendance recorded in: `{att_path}`")
                    else:
                        st.warning("No faces were captured.")
